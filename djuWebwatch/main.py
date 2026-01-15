# 웹접근성 검사를 위해 게시물을 검사하는 프로그램
# 검사하는 내용
# img 태그가 있는 경우 alt를 검색
# table 태그가 있는 경우 caption을 검색
# a 태그에 target="_blank" 인 경우 title을 검색

from urllib.parse import urlparse, parse_qs
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Alignment
from datetime import datetime
import os
import unicodedata

# url을 조회한다
def get_query(url):
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.82 Safari/537.36'
        }
        response = requests.get(url, headers=headers)
        response.raise_for_status()  # 요청이 성공했는지 확인
        soup = BeautifulSoup(response.text, 'html.parser')
        return soup
    except Exception as e:
        print(f'오류 발생: {e}')
        
    return None

# 게시물 url을 수집한다
def get_ntt_url_list(bbs_url_list, max_page):
    ntt_url_list = []
    for bbs_url in bbs_url_list:
        parsed_url = urlparse(bbs_url)
        query_string = parsed_url.query
        params = parse_qs(query_string)
        mi = params.get('mi')[0]
        bbsId = params.get('bbsId')[0]
        
        for page in range(1, max_page + 1):
            url = bbs_url + f'&currPage={page}'
            print(f'조회 : {url}')
            soup = get_query(url)
            links = soup.select('.nttInfoBtn')
            for link in links:
                ntt_url_list.append(f'https://www.dju.ac.kr/dju/na/ntt/selectNttInfo.do?nttSn={link.get('data-id')}&bbsId={bbsId}&mi={mi}')
    
    return ntt_url_list

# img, table, a 태그를 검사한다
def search_ntt_contents(ntt_url):
    soup = get_query(ntt_url)
    
    contents = soup.select_one('.BD_table table .Cnts')
    imgs = contents.select('img')
    tables = contents.select('table')
    a = contents.select('a[target="_blank"]')

    if not imgs and not tables and not a:
        return None
    
    desc = []

    for img in imgs:
        desc.append(f'src="{img.get('src')}" : {'alt="'+img.get('alt')+'"' if img.get('alt') else 'alt 없음'}')

    desc.extend(a)

    for table in tables:
        caption = table.find('caption')
        if caption is None:
            desc.append('caption 없는 테이블')
        else: 
            desc.append(caption)
    for i in range(len(desc)):
        desc[i] = str(desc[i])
    return [ntt_url, '\n'.join(desc)]

# 텍스트의 "유효 너비"를 계산하는 함수
def get_effective_width(text):
    if text is None:
        return 0
    
    text_str = str(text) # 숫자가 들어올 수도 있으니 일단 문자열로 변환
    effective_width = 0
    
    for char in text_str:
        # unicodedata.east_asian_width()는 문자의 폭을 'F'(전각), 'W'(전각), 'A'(애매), 'H'(반각), 'Na'(좁음) 등으로 반환해.
        # 한글, 한자 등은 주로 'F' 또는 'W' (전각), 영문/숫자는 'Na' 또는 'H' (반각)
        if unicodedata.east_asian_width(char) in ('F', 'W', 'A'):
            effective_width += 2  # 전각 문자는 너비를 2로 간주 (한글, 한자)
        else:
            effective_width += 1  # 반각 문자는 너비를 1로 간주 (영문, 숫자, 대부분의 특수문자)
            
    return effective_width

# 엑셀 열 넓이 조절
def excel_col_auto_size(sheet, col):
    # 길이 조절
    cls_column = sheet[f'{col}:{col}']  # B column의 width를 autofit할 것이기 때문에 B column 객체 얻어옴
    max_length = 0
    for cell in cls_column:  # B column에 있는 cell을 하나씩 참조
        try:  # blank cell에서 값을 참조하려면 error가 발생하기 때문에 try ~ excep 구문 사용
            splited_str = str(cell.value).split('\n')
            for s in splited_str:
                width = get_effective_width(s)
                if width > max_length:
                    max_length = width
        except:
            pass
    sheet.column_dimensions['B'].width = max_length

# 엑셀 파일로 저장
def save_to_excel(ntt_content_list):
    # 1. 새 워크북 생성
    workbook = openpyxl.Workbook()

    # 2. 현재 활성화된 시트 선택 (기본적으로 첫 번째 시트)
    sheet = workbook.active
    sheet.title = '대전대 웹접근성' # 시트 이름 설정

    # 3. 헤더 (컬럼명) 작성
    headers = ['URL', '확인 필요 내용']
    sheet.append(headers)

    # 사이즈 조절
    sheet.column_dimensions['A'].width = 76

    # 4. 데이터 추가
    for row_data in  ntt_content_list:
        sheet.append(row_data)
        # 방금 추가된 행의 번호 (sheet.max_row는 현재 시트의 마지막 행 번호를 반환)
        current_row_idx = sheet.max_row

        # B 컬럼(두 번째 컬럼, 인덱스 2)의 셀에 접근
        b_column_cell = sheet.cell(row=current_row_idx, column=2)

        # 해당 셀에 '텍스트 줄 바꿈' 설정 (이게 핵심!)
        b_column_cell.alignment = Alignment(wrap_text=True)

    excel_col_auto_size(sheet, 'B')

    current_time = datetime.now()
    formatted_time_string = current_time.strftime('%y%m%d%H%M%S')

    # 5. 엑셀 파일 저장
    os.makedirs('./결과', exist_ok=True)
    file_name = f'./결과/대전대 웹접근성_{formatted_time_string}.xlsx'
    workbook.save(file_name)

    print(f'"{file_name}" 파일이 성공적으로 저장되었습니다.')

def main():
    bbs_url_list = []

    while True:
        url = input('게시판의 URL 입력 : ')
        if url == '':
            break
        bbs_url_list.append(url)
    
    # 입력한 게시판에서 조회해야하는 게시물 url을 수집한다
    if bbs_url_list.count == 0:
        print('입력한 URL이 없음')
        return
    
    max_page = int(input('최대 페이지 수를 입력 : '))
    print('게시물 url 수집 중...')
    ntt_url_list = get_ntt_url_list(bbs_url_list, max_page)
    print('게시물 url 수집 완료')
    ntt_url_list = list(set(ntt_url_list))
    # 게시물 돌아다니면서 내용을 검출하고 검출된 내용을 저장한다
    ntt_content_list = []
    curr_progress = 1
    max_progress = len(ntt_url_list)
    print('게시물 검출 중...')
    for ntt_url in ntt_url_list:
        print(f'[{curr_progress}/{max_progress}] {ntt_url}')
        curr_progress += 1
        contents = search_ntt_contents(ntt_url)
        if contents is not None:
            ntt_content_list.append(contents)
    print('게시물 검출 완료')
    # 엑셀로 게시물 url, 검출 내용을 저장한다
    save_to_excel(ntt_content_list)

if __name__ == '__main__':
    main()
    input('Enter key to exit...')